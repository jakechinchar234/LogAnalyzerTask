
"""
BitSwitchingDetector - Packetswitch component extractor
Author: Jacob King
Last Edited: 2025-12-05

Purpose
-------
Parse a Packetswitch HTML/DOCX file and an Excel file containing
signal labels to produce a  summary of bit changes.
Returns a dictionary keyed by (time_tag, type) where type is 'Ind' or 'Control'.
"""

import re
from typing import Dict, List, Tuple
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from docx import Document

# ---------- File readers ----------

def read_packetswitch_lines(file_path: str) -> List[str]:
    """Read Packetswitch file (.html or .docx) and return raw lines of text."""
    p = Path(file_path)
    if not p.exists():
        return []
    if p.suffix.lower() == ".html":
        with open(p, "r", encoding="utf-8", errors="ignore") as f:
            soup = BeautifulSoup(f, "html.parser")
            text = soup.get_text()
        return text.strip().split("\n")
    elif p.suffix.lower() == ".docx":
        try:
            doc = Document(str(p))
            text = "\n".join([para.text for para in doc.paragraphs])
        except Exception:
            # fallback: try reading as text
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        return text.strip().split("\n")
    else:
        # generic text fallback
        with open(p, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()
        return text.strip().split("\n")


def read_signal_labels(excel_path: str) -> Tuple[List[str], List[str]]:
    """Read Indication and Control labels from Column D of two sheets."""
    wb = load_workbook(filename=excel_path, data_only=True)
    names = wb.sheetnames

    # Prefer sheets whose titles contain 'ind' and 'control'
    ind_sheet = None
    ctrl_sheet = None
    for name in names:
        title = wb[name].title.lower()
        if ind_sheet is None and ("ind" in title or "indic" in title):
            ind_sheet = wb[name]
        if ctrl_sheet is None and ("control" in title or "ctrl" in title):
            ctrl_sheet = wb[name]
    # Fallback: take first two sheets if detection failed
    if ind_sheet is None or ctrl_sheet is None:
        if len(names) >= 2:
            s0, s1 = wb[names[0]], wb[names[1]]
            if "control" in s0.title.lower():
                ctrl_sheet, ind_sheet = s0, s1
            else:
                ind_sheet, ctrl_sheet = s0, s1
        else:
            # Not enough sheets
            return [], []

    def col_d_values(ws):
        labels: List[str] = []
        for col in ws.iter_cols(min_row=2, min_col=4, max_col=4, values_only=True):
            for v in col:
                if v is None:
                    continue
                labels.append(str(v).strip())
        return labels

    indications = col_d_values(ind_sheet)
    controls    = col_d_values(ctrl_sheet)
    return indications, controls

# ---------- Core logic ----------

def _hex_list_to_binary_string(hex_list: List[str]) -> str:
    """Convert list of hex byte strings to a concatenated 8-bit binary string."""
    bits = []
    for h in hex_list:
        try:
            b = format(int(h, 16), "08b")
            bits.append(b)
        except Exception:
            # skip non-hex tokens
            continue
    return "".join(bits)


def detect_bit_switches(ps_lines: List[str], indications: List[str], controls: List[str]) -> List[Dict[str, str]]:
    """
    Process Packetswitch lines, derive On/Off transitions per timestamp and type.
    Returns list of dicts: { 'time': HH:MM:SS, 'type': 'Ind'|'Control', 'component': '...'}
    """
    IndicPrev: List[str] = []
    CtrlPrev: List[str] = []
    results: List[Dict[str, str]] = []

    for i, line in enumerate(ps_lines, start=1):
        # Lines like: "P# HH:MM:SS Type ... : <hex hex ...>"
        if ": " not in line:
            continue
        head, data_hex = line.split(": ", 1)
        head = head.strip()
        data_hex = data_hex.strip()

        # Skip "Indicate" (non-RF) lines, matching your original logic
        if "Indicate" in head:
            continue

        tokens = [t for t in head.split(" ") if t != ""]
        if len(tokens) < 3:
            continue

        # Timestamp expected at tokens[1]
        timestamp = tokens[1] if re.match(r"\d{2}:\d{2}:\d{2}", tokens[1]) else None
        if not timestamp:
            continue

        tkns_lower = [t.lower() for t in tokens]
        if any("indic" in t for t in tkns_lower):
            comm_type = "Ind"
        elif any("control" in t for t in tkns_lower):
            comm_type = "Control"
        else:
            # Ignore Recall / other types
            continue

        hex_items = [h for h in data_hex.split(" ") if re.fullmatch(r"[0-9A-Fa-f]{2}", h)]
        if not hex_items:
            continue
        status_bits = _hex_list_to_binary_string(hex_items)

        # Map bits → signal names
        active: List[str] = []
        if comm_type == "Ind":
            for idx, bit in enumerate(status_bits):
                if bit == "1" and idx < len(indications):
                    active.append(indications[idx])
        else:  # Control
            for idx, bit in enumerate(status_bits):
                if bit == "1" and idx < len(controls):
                    active.append(controls[idx])

        # Changes vs previous state
        changed = []
        if comm_type == "Ind":
            for s in active:
                if s not in IndicPrev:
                    changed.append(f"{s} On")
            for s in IndicPrev:
                if s not in active:
                    changed.append(f"{s} Off")
        else:
            for s in active:
                if s not in CtrlPrev:
                    changed.append(f"{s} On")
            for s in CtrlPrev:
                if s not in active:
                    changed.append(f"{s} Off")

        if changed:
            results.append({
                "time": timestamp,
                "type": comm_type,
                "component": " / ".join(changed)
            })

        # Update previous state
        if comm_type == "Ind":
            IndicPrev = active[:]
        else:
            CtrlPrev = active[:]

    return results

# ---------- Public API ----------

def build_component_map(packetswitch_file_path: str, excel_file_path: str) -> Dict[Tuple[str, str], str]:
    """Read files and return {(HH:MM:SS, 'Ind'|'Control'): component}."""
    ps_lines = read_packetswitch_lines(packetswitch_file_path)
    if not ps_lines:
        return {}
    indications, controls = read_signal_labels(excel_file_path)
    if not indications and not controls:
        return {}
    entries = detect_bit_switches(ps_lines, indications, controls)
    comp_map: Dict[Tuple[str, str], str] = {}
    for e in entries:
        comp_map[(e["time"], e["type"])] = e["component"].strip()
    return comp_map
